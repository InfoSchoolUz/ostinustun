"""
AI Leaders PINFL Checker — To'liq avtomatik
============================================
requirements.txt:
    streamlit
    openpyxl
    pandas
    requests
"""
import re
import time
import requests
import pandas as pd
import streamlit as st
from io import BytesIO
from openpyxl.styles import numbers as xl_numbers

# ──────────────────────────────────────────
# CONFIG
# ──────────────────────────────────────────
API_URL    = "https://aileaders.uz/api/v1/check/certificates"
DELAY_SEC  = 1.5   # rate limit uchun

# ──────────────────────────────────────────
# cURL DAN COOKIE AJRATISH
# ──────────────────────────────────────────
def parse_curl(curl_text: str) -> dict:
    result = {
        "cookie": "", "user_agent": "",
        "accept": "*/*", "accept_language": "",
        "referer": "", "extra_headers": {}
    }

    # Cookie — -b '...' yoki -H 'cookie: ...'
    for pattern in [
        r"-b\s+'([^']+)'", r'-b\s+"([^"]+)"',
        r"-H\s+'cookie:\s*([^']+)'", r'-H\s+"cookie:\s*([^"]+)"',
    ]:
        m = re.search(pattern, curl_text, re.IGNORECASE)
        if m:
            result["cookie"] = m.group(1).strip()
            break

    # Barcha -H headerlarini olish
    for m in re.finditer(r"-H\s+'([^']+)'|-H\s+\"([^\"]+)\"", curl_text):
        raw = (m.group(1) or m.group(2)).strip()
        if ":" not in raw:
            continue
        key, _, val = raw.partition(":")
        key = key.strip().lower()
        val = val.strip()
        if key == "user-agent":
            result["user_agent"] = val
        elif key == "accept" and "language" not in key:
            result["accept"] = val
        elif key == "accept-language":
            result["accept_language"] = val
        elif key == "referer":
            result["referer"] = val
        elif key not in ("cookie", "content-length"):
            result["extra_headers"][key] = val

    return result

# ──────────────────────────────────────────
# EXCEL O'QISH — original ustunlar saqlanadi
# ──────────────────────────────────────────
def read_excel(file) -> tuple[pd.DataFrame, str, str]:
    """
    Original DataFrame, PINFL ustun nomi, sana ustun nomini qaytaradi.
    Header qaysi qatorda bo'lishidan qat'i nazar topadi.
    """
    xls = pd.ExcelFile(file, engine="openpyxl")
    sheet_name = xls.sheet_names[0]

    # Header qatorini qidirish (0-7 qator orasida)
    header_row = 0
    for i in range(8):
        try:
            df_tmp = pd.read_excel(xls, sheet_name=sheet_name, header=i, nrows=1, engine="openpyxl")
            cols = df_tmp.columns.astype(str).str.upper()
            if cols.str.contains(r"ПИНФЛ|PINFL", regex=True).any():
                header_row = i
                break
        except Exception:
            continue

    df = pd.read_excel(xls, sheet_name=sheet_name, header=header_row, engine="openpyxl")
    df.columns = df.columns.map(str).str.strip()

    # PINFL ustuni
    pinfl_col = next(
        (c for c in df.columns if "ПИНФЛ" in c.upper() or "PINFL" in c.upper()), None
    )
    if pinfl_col is None:
        return pd.DataFrame(), "", ""

    # Sana ustuni
    date_col = next(
        (c for c in df.columns
         if "рождени" in c.lower() or "birth" in c.lower() or "sana" in c.lower()),
        None
    )

    # Bo'sh qatorlarni tashla
    df = df[df[pinfl_col].notna()].copy()

    # PINFL — matn sifatida, .0 ni olib tashla, 14 xonaga to'ldirish
    df[pinfl_col] = (
        df[pinfl_col].astype(str).str.strip()
        .str.replace(r"\.0$", "", regex=True)
        .str.replace(r"\s+", "", regex=True)
    )
    df = df[df[pinfl_col].str.len() >= 10].reset_index(drop=True)

    # Sana — datetime formatda saqlash
    if date_col:
        df[date_col] = pd.to_datetime(df[date_col], errors="coerce")

    return df, pinfl_col, date_col or ""

# ──────────────────────────────────────────
# BITTA PINFL TEKSHIRISH
# ──────────────────────────────────────────
def check_pinfl(pinfl: str, session: requests.Session) -> dict:
    empty = {"holat": "🔴 Xato", "ism": "", "email": "", "kurslar": [], "xato": ""}
    try:
        r = session.get(API_URL, params={"pinfl": pinfl}, timeout=20)

        if r.status_code in (200, 304):
            try:
                data = r.json()
            except Exception:
                return {**empty, "holat": "⚠️ JSON xato", "xato": r.text[:80]}

            all_courses = data.get("courses", [])
            completed   = [c for c in all_courses if c.get("isCompleted")]

            import datetime as _dt
            kurslar = []
            for c in completed:
                url = c.get("contentCertificateUrl") or ""

                raw_ts = c.get("completedAt") or ""
                if str(raw_ts).isdigit() and int(raw_ts) > 1_000_000_000:
                    ts = int(raw_ts)
                    if ts > 1e10:
                        ts = ts / 1000
                    vaqt = _dt.datetime.fromtimestamp(ts).strftime("%d.%m.%Y")
                else:
                    vaqt = str(raw_ts)

                kurslar.append({
                    "nomi": c.get("contentName", "").strip(),
                    "url":  url,
                    "vaqt": vaqt,
                })

            has_courses = data.get("hasCourses", False)
            return {
                "holat":   "✅ Sertifikat bor" if kurslar else ("⚠️ Ro'yxatda bor, kurs yo'q" if has_courses else "⚠️ Kurs bor, sertifikat yo'q"),
                "ism":     data.get("fullName", ""),
                "email":   data.get("email", ""),
                "kurslar": kurslar,
                "xato":    "",
            }

        elif r.status_code == 404:
            return {**empty, "holat": "❌ Topilmadi",       "kurslar": []}
        elif r.status_code == 401:
            return {**empty, "holat": "🔐 Cookie eskirgan", "kurslar": [], "xato": "Cookie yangilang"}
        elif r.status_code == 429:
            time.sleep(60)
            return {**empty, "holat": "⏳ Rate limit",      "kurslar": [], "xato": "1 daqiqa kutildi"}
        else:
            return {**empty, "holat": f"🔴 {r.status_code}", "kurslar": [], "xato": r.text[:80]}

    except Exception as e:
        return {**empty, "holat": "🔴 Xato", "kurslar": [], "xato": str(e)[:80]}

# ──────────────────────────────────────────
# EXCEL EKSPORT — formatlar saqlanadi
# ──────────────────────────────────────────
def export_excel(result_df: pd.DataFrame, pinfl_col: str, date_col: str) -> bytes:
    out = BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        result_df.to_excel(writer, index=False, sheet_name="Natijalar")
        ws = writer.sheets["Natijalar"]

        # 1-qator — ustun nomlari (header)
        header_map = {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1)}

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                col_name = str(ws.cell(1, cell.column).value or "")

                # PINFL — matn (scientific notation bo'lmasin)
                if "ПИНФЛ" in col_name.upper() or "PINFL" in col_name.upper():
                    cell.number_format = "@"
                    if cell.value is not None:
                        cell.value = str(cell.value).replace(".0", "").strip()

                # Tug'ilgan sana
                elif date_col and col_name == date_col:
                    cell.number_format = "DD.MM.YYYY"

                # ID, raqamli ustunlar (Серия, Номер документа)
                elif any(k in col_name for k in ["Серия", "Номер", "ID", "№"]):
                    cell.number_format = "@"
                    if cell.value is not None:
                        cell.value = str(cell.value).replace(".0", "").strip()

                # URL — matn
                elif "URL" in col_name.upper():
                    cell.number_format = "@"

                # Kurs vaqti — matn
                elif "vaqti" in col_name.lower():
                    cell.number_format = "@"

    return out.getvalue()

# ──────────────────────────────────────────
# STREAMLIT UI
# ──────────────────────────────────────────
st.set_page_config(
    page_title="AI Leaders PINFL Checker",
    page_icon="🎓",
    layout="wide"
)
st.title("🎓 AI Leaders — PINFL Sertifikat Tekshiruvi")
st.caption("Excel fayldagi barcha PINFLlarni aileaders.uz da avtomatik tekshiradi va natijani original formatda qaytaradi")

# ── 1-QADAM: cURL ──
st.subheader("🔐 1-qadam: cURL joylashtiring")
with st.expander("📋 cURL qanday olish kerak?", expanded=True):
    st.markdown("""
    1. **`https://aileaders.uz/auth/login/check`** saytini oching
    2. Istalgan PINFL kiriting → **Tekshirish** bosing
    3. **F12** → **Network** tab
    4. `certificates?pinfl=...` qatoriga **o'ng klik**
    5. **Copy → Copy as cURL (bash)** tanlang
    6. Quyidagi maydonga **Ctrl+V**
    """)

curl_input = st.text_area(
    "cURL matni:",
    placeholder="curl 'https://aileaders.uz/api/v1/check/certificates?pinfl=...' \\\n  -H 'cookie: HWWAFSESID=...' \\\n  ...",
    height=120,
)

parsed  = {}
curl_ok = False
if curl_input.strip():
    parsed = parse_curl(curl_input)
    if parsed["cookie"] and "HWWAFSESID" in parsed["cookie"]:
        st.success("✅ Cookie muvaffaqiyatli topildi!")
        curl_ok = True
    else:
        st.error("❌ Cookie topilmadi. cURL to'liq ko'chirilganmi?")

st.divider()

# ── 2-QADAM: EXCEL ──
st.subheader("📂 2-qadam: Excel yuklang")
uploaded = st.file_uploader(
    "O'quvchilar ro'yxati (xlsx)",
    type=["xlsx"],
    label_visibility="collapsed"
)

if uploaded:
    with st.spinner("Excel o'qilmoqda..."):
        df_orig, pinfl_col, date_col = read_excel(uploaded)

    if df_orig.empty:
        st.error("❌ PINFL ustuni topilmadi! Ustun nomi ПИНФЛ yoki PINFL bo'lishi kerak.")
        st.stop()

    st.success(f"✅ **{len(df_orig)}** ta o'quvchi topildi | PINFL ustuni: **{pinfl_col}**")

    # Preview
    with st.expander("📋 Ma'lumotlarni ko'rish (dastlabki 5 ta)"):
        st.dataframe(df_orig.head(), use_container_width=True)

    # Taxminiy vaqt
    daqiqa = round(len(df_orig) * DELAY_SEC / 60, 1)
    st.info(f"⏱️ Taxminiy vaqt: **{daqiqa} daqiqa** ({len(df_orig)} ta × {DELAY_SEC}s)")

    st.divider()

    # ── 3-QADAM: TEKSHIRISH ──
    st.subheader("🚀 3-qadam: Tekshirishni boshlash")

    if not curl_ok:
        st.warning("⚠️ Avval 1-qadamda cURL ni joylashtiring!")

    if curl_ok and st.button("🚀 Tekshirishni boshlash", type="primary", use_container_width=True):

        session = requests.Session()
        session.headers.update({
            "User-Agent":      parsed.get("user_agent") or (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/148.0.0.0 Safari/537.36"
            ),
            "Accept":          parsed.get("accept", "*/*"),
            "Accept-Language": parsed.get("accept_language", "ru,en-US;q=0.9,en;q=0.8"),
            "Referer":         parsed.get("referer", "https://aileaders.uz/auth/login/check"),
            "Cookie":          parsed["cookie"],
            **parsed.get("extra_headers", {}),
        })

        progress_bar = st.progress(0)
        status_text  = st.empty()
        total        = len(df_orig)
        all_kurslar  = []
        cookie_dead  = False
        # ✅ Loop DAN OLDIN yaratiladi
        result_df = df_orig.copy()
        result_df["Email"] = ""
        result_df["Holat"] = ""

        for i, row in df_orig.iterrows():
            pinfl = str(row[pinfl_col])
            fio   = str(row.get("Полное наименование", row.get("F.I.Sh.", "")))
            status_text.text(f"🔄 {i+1}/{total} — {pinfl} | {fio[:40]}")
            progress_bar.progress((i + 1) / total)

            res = check_pinfl(pinfl, session)

            if res["holat"] == "🔐 Cookie eskirgan":
                st.error("🔐 Cookie eskirdi! Yangi cURL olib, qayta boshlang.")
                cookie_dead = True
                # Qolgan qatorlar uchun bo'sh yozuv
                for _ in range(total - i - 1):
                    all_kurslar.append([])
                all_kurslar.insert(i, [])
                break

            all_kurslar.append(res["kurslar"])
            result_df.at[i, "Email"] = res["email"]
            result_df.at[i, "Holat"] = res["holat"]
            time.sleep(DELAY_SEC)

        if not cookie_dead:
            status_text.success(f"✅ Yakunlandi! {len(all_kurslar)}/{total} ta tekshirildi")
        progress_bar.progress(1.0)

        # ── Dinamik ustunlar qo'shish ──
        max_kurs = max((len(k) for k in all_kurslar), default=0)
        MAX_KURS = 10  # Har doim 10 ta kurs uchun ustun ochiladi

        for idx, kurslar in enumerate(all_kurslar):
            for n in range(MAX_KURS):
                if n < len(kurslar):
                    result_df.at[idx, f"Kurs {n+1} nomi"] = kurslar[n]["nomi"]
                    result_df.at[idx, f"Kurs {n+1} URL"]  = kurslar[n]["url"]
                    result_df.at[idx, f"Kurs {n+1} vaqti"]= kurslar[n]["vaqt"]
                else:
                    result_df.at[idx, f"Kurs {n+1} nomi"] = ""
                    result_df.at[idx, f"Kurs {n+1} URL"]  = ""
                    result_df.at[idx, f"Kurs {n+1} vaqti"]= ""

        # ── Statistika ──
        sertifikat_bor = sum(1 for k in all_kurslar if k)
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Jami tekshirildi",  total)
        c2.metric("✅ Sertifikat bor", sertifikat_bor)
        c3.metric("❌ Yo'q/Topilmadi", total - sertifikat_bor)
        c4.metric("📚 Maks kurslar",   max_kurs)

        st.subheader("📋 Natija")
        st.dataframe(result_df, use_container_width=True)

        # ── Excel eksport ──
        excel_bytes = export_excel(result_df, pinfl_col, date_col)
        st.download_button(
            "📥 Natijani yuklab olish (Excel)",
            data=excel_bytes,
            file_name="aileaders_natija.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

st.markdown(
    '<div style="position:fixed;left:0;bottom:0;width:100%;background:#0e1117;'
    'color:white;text-align:center;padding:8px;font-weight:bold;z-index:1000;">'
    'Tuzuvchi: Azamat Madrimov | 2026</div>',
    unsafe_allow_html=True
)
